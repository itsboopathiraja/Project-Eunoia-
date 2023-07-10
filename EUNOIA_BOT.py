import telebot
from openpyxl import Workbook, load_workbook
import re
import random
from telebot import types
import time

bot = telebot.TeleBot('6226498182:AAEkAVpmliSt0UFZMswazB2ZhjpWAH07hhk')
filename = 'user_data.xlsx'
wb = Workbook()
ws = wb.active
ws.append(['Name', 'Age', 'DOB', 'Register Number', 'ID Card', 'Department', 'Year', 'Contact Number', 'WhatsApp Number', 'Email', 'Father Name', 'Mother Name', 'Parents Number', 'Father Work', 'Mother Work', '10th Mark (out of 500)', '12th Mark (out of 600)', 'School Name (10th)', 'School Name (12th)', 'Address'])
wb.save(filename)
faculty_password = '9933'
user_password = 'shanmugha'

# Create a custom keyboard
keyboard = types.ReplyKeyboardMarkup(row_width=2)
# Add buttons for each command
keyboard.add('Students', 'Faculty')

# Create inline keyboards for student, faculty, and admin commands
student_keyboard = types.InlineKeyboardMarkup()
student_keyboard.add(types.InlineKeyboardButton("Enter Details", callback_data="enterdetails"))
student_keyboard.add(types.InlineKeyboardButton("Edit Details", callback_data="editdetails"))
student_keyboard.add(types.InlineKeyboardButton("Delete Detail", callback_data="delete_detail"))
student_keyboard.add(types.InlineKeyboardButton("newupdates", callback_data="newupdates"))

faculty_keyboard = types.InlineKeyboardMarkup()
faculty_keyboard.add(types.InlineKeyboardButton("Search Details", callback_data="searchdetails"))
faculty_keyboard.add(types.InlineKeyboardButton("Get File", callback_data="get_file"))
faculty_keyboard.add(types.InlineKeyboardButton("Reset Sheet", callback_data="resetsheet"))
faculty_keyboard.add(types.InlineKeyboardButton("Delete Detail", callback_data="delete_detail"))
faculty_keyboard.add(types.InlineKeyboardButton("newupdates", callback_data="newupdates"))

@bot.message_handler(func=lambda message: message.text == 'Students')
def handle_students_button(message):
    bot.send_message(message.chat.id, "Here are the available student commands:", reply_markup=student_keyboard)

@bot.message_handler(func=lambda message: message.text == 'Faculty')
def handle_faculty_button(message):
    bot.send_message(message.chat.id, "Here are the available faculty commands:", reply_markup=faculty_keyboard)

@bot.callback_query_handler(func=lambda call: True)
def handle_callback_query(call):
    # Send a message to the user immediately after they click an inline button
    bot.answer_callback_query(call.id, text="Processing your request...")

    if call.data == "enterdetails":
        # Handle enterdetails command
        enterdetails(call.message)
    elif call.data == "editdetails":
        # Handle editdetails command
        editdetails(call.message)
    elif call.data == "searchdetails":
        # Handle searchdetails command
        search_details(call.message)
    elif call.data == "get_file":
        # Handle get_file command
        get_file(call.message)
    elif call.data == "resetsheet":
        # Handle resetsheet command
        resetsheet(call.message)
    elif call.data == "delete_detail":
        # Handle delete_detail command
        delete_detail(call.message)
    elif call.data == "newupdates":
        # Handle delete_detail command
        newupdates(call.message)
    elif call.data == "add_admin":
        # Handle add_admin command
        add_admin(call.message)
    elif call.data == "remove_admin":
        # Handle remove_admin command
        remove_admin(call.message)
    elif call.data == "update":
        # Handle remove_admin command
        update(call.message)
    elif call.data == "deleteupdate":
        # Handle remove_admin command
        deleteupdate(call.message)


videos = ['/home/boopathiraj/videos/video_1.mp4', '/home/boopathiraj/videos/video_2.mp4', '/home/boopathiraj/videos/video_3.mp4']

restricted_commands = ["""
- /Help_Me\n
- /how_it_works\n
- /about\n
- /feedback\n
"""]

@bot.message_handler(commands=['start'])
def send_welcome(message):
    video_path = random.choice(videos)
    video = open(video_path, 'rb')
    bot.send_video(message.chat.id, video, caption="Welcome! I'm excited to be working with you. Let's make some magic happen.", reply_markup=keyboard)
    time.sleep(3)
    msg = bot.send_message(message.chat.id, "Which type of user are you? select one\n\n     1. /Just_Explore         2. /SSEI")
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
    markup.add('Just Explore', 'SSEI')
    bot.register_next_step_handler(msg, check_user_type, markup)

def check_user_type(message, markup):
    if message.text == '/Just_Explore':
        bot.send_message(message.chat.id, "You have restricted access to this bot. You can only use the following commands:\n {}".format(', '.join(restricted_commands)))
    elif message.text == '/SSEI':
        msg = bot.send_message(message.chat.id, "Please enter your user password to access me.")
        bot.register_next_step_handler(msg, check_password)

def check_password(message):
    if message.text == user_password:
        bot.send_message(message.chat.id, "Happy to help you! Here are the available commands:")
        time.sleep(2)
        send_help(message)
    else:
        bot.send_message(message.chat.id, "Sorry, this bot is only for Shanmugha college students.")



import os
import json
import time

admin_id = '6293791327'
admin_ids = [admin_id]
updates_file = 'updates.json'

admin_keyboard = types.ReplyKeyboardMarkup(row_width=2)
admin_keyboard.add('/add_admin', '/remove_admin', '/update', '/deleteupdate')

keyboard.add('/admin')

@bot.message_handler(commands=['admin'])
def admin(message):
    if str(message.from_user.id) in admin_ids:
        bot.send_message(message.chat.id, "Here are the available admin commands:", reply_markup=admin_keyboard)
    else:
        bot.send_message(message.chat.id, "Sorry, this option is only for admins.")


@bot.message_handler(commands=['add_admin'])
def add_admin(message):
    if str(message.from_user.id) in admin_ids:
        msg = bot.send_message(message.chat.id, "Please enter the user ID of the user you want to add as an admin.")
        bot.register_next_step_handler(msg, add_admin_step)
    else:
        bot.send_message(message.chat.id, "Only admins can use this command.")

def add_admin_step(message):
    user_id = message.text
    if user_id not in admin_ids:
        admin_ids.append(user_id)
        bot.send_message(message.chat.id, f"User {user_id} has been added as an admin.")
    else:
        bot.send_message(message.chat.id, f"User {user_id} is already an admin.")


@bot.message_handler(commands=['remove_admin'])
def remove_admin(message):
    if str(message.from_user.id) in admin_ids:
        msg = bot.send_message(message.chat.id, "Please enter the user ID of the user you want to remove as an admin.")
        bot.register_next_step_handler(msg, remove_admin_step)
    else:
        bot.send_message(message.chat.id, "Only admins can use this command.")

def remove_admin_step(message):
    user_id = message.text
    if user_id in admin_ids:
        admin_ids.remove(user_id)
        bot.send_message(message.chat.id, f"User {user_id} has been removed as an admin.")
    else:
        bot.send_message(message.chat.id, f"User {user_id} is not an admin.")


if os.path.exists(updates_file):
    with open(updates_file, 'r') as f:
        updates = json.load(f)
else:
    updates = []

@bot.message_handler(commands=['update'])
def update(message):
    if str(message.from_user.id) in admin_ids:
        msg = bot.send_message(message.chat.id, "Please enter the new updates.")
        bot.register_next_step_handler(msg, check_media)
    else:
        bot.send_message(message.chat.id, "Only admins can use this command.")


def check_media(message):
    if message.content_type in ['photo', 'video', 'document']:
        handle_media(message)
    else:
        save_update(message)

def save_update(message):
    if message.text:
        update = {'type': 'text', 'content': message.text, 'timestamp': time.time()}
        updates.append(update)
        with open(updates_file, 'w') as f:
            json.dump(updates, f)
        bot.send_message(message.chat.id, "Update saved.")
    else:
        bot.send_message(message.chat.id, "Please enter a non-empty update.")

@bot.message_handler(content_types=['photo', 'video', 'document'])
def handle_media(message):
    if str(message.from_user.id) == admin_id:
        file_id = None
        file_type = None
        caption = message.caption # get the caption of the media file
        if message.photo:
            file_id = message.photo[-1].file_id
            file_type = 'photo'
        elif message.video:
            file_id = message.video.file_id
            file_type = 'video'
        elif message.document:
            file_id = message.document.file_id
            file_type = 'document'
        if file_id:
            update = {'type': file_type, 'content': file_id, 'caption': caption, 'timestamp': time.time()}
            updates.append(update)
            with open(updates_file, 'w') as f:
                json.dump(updates, f)
            bot.send_message(message.chat.id, f"{file_type.capitalize()} update saved.")


@bot.message_handler(commands=['newupdates'])
def newupdates(message):
    current_time = time.time()
    has_updates = False # flag to check if there are any new updates
    for update in updates[:]:
        if current_time - update['timestamp'] < 3 * 24 * 60 * 60: # check if update is less than 3 days old
            has_updates = True # set flag to True if there is at least one new update
            if update['type'] == 'text' and update['content']:
                bot.send_message(message.chat.id, update['content'])
            elif update['type'] in ['photo', 'video', 'document']:
                send_func = getattr(bot, f"send_{update['type']}")
                send_func(message.chat.id, update['content'], caption=update.get('caption')) # send media file with its caption
        else: # delete update if it's more than 3 days old
            updates.remove(update)
            with open(updates_file, 'w') as f:
                json.dump(updates, f)
    if not has_updates: # if there are no new updates
        bot.send_message(message.chat.id, "There are no new updates.")

@bot.message_handler(commands=['deleteupdate'])
def deleteupdate(message):
    if str(message.from_user.id) == admin_id:
        updates.clear() # clear the updates list
        with open(updates_file, 'w') as f:
            json.dump(updates, f)
        bot.send_message(message.chat.id, "All updates deleted.")
    else:
        bot.send_message(message.chat.id, "Only the admin can use this command.")


entering_details = False
waiting_users = []

@bot.message_handler(commands=['enterdetails'])
def enterdetails(message):
    global entering_details
    if entering_details:
        bot.send_message(message.chat.id, "Please wait, another user is currently entering their details.")
        waiting_users.append(message.chat.id)
    else:
        entering_details = True
        bot.send_message(message.chat.id, "Please note that all details entered into this bot will be stored and accessible by faculty. Kindly ask that you take responsibility for the information you provide and answer all questions truthfully.")
        time.sleep(5)
        bot.send_message(message.chat.id, "Hello! Please enter your name.")
        bot.register_next_step_handler(message, get_name)


def notify_waiting_users():
    global waiting_users
    for chat_id in waiting_users:
        bot.send_message(chat_id, "The bot is now free to take your details. Please enter the /enterdetails command again.")
    waiting_users = []

def get_name(message):
    name = message.text
    if re.match(r'^[A-Z]\. [a-zA-Z\s]+$', name):
        ws.append([name])
        bot.send_message(message.chat.id, f"👋 Hello {name}! How old are you?")
        bot.register_next_step_handler(message, get_age, name)
    else:
        bot.send_message(message.chat.id, "🤔 Hmm, that doesn't look quite right. Could you please enter your name in the format 'Initial. Name' (e.g. 'K. Ram')?")
        bot.register_next_step_handler(message, get_name)

def get_age(message, name):
    age = message.text
    if age.isdigit():
        ws.cell(row=ws.max_row, column=2).value = age
        bot.send_message(message.chat.id, f"🎂 Got it! And when is your birthday, {name}? Please enter it in the format DD/MM/YYYY.")
        bot.register_next_step_handler(message, get_dob, name)
    else:
        bot.send_message(message.chat.id, "🤔 That doesn't look like a number. Could you please enter your age using only digits?")
        bot.register_next_step_handler(message, get_age, name)

def get_dob(message, name):
    dob = message.text
    if re.match(r'^\d{2}/\d{2}/\d{4}$', dob):
        ws.cell(row=ws.max_row, column=3).value = dob
        bot.send_message(message.chat.id, f"📅 Great! And what is your register number, {name}?")
        bot.register_next_step_handler(message, get_register_number, name)
    else:
        bot.send_message(message.chat.id, "🤔 That doesn't look quite right. Could you please enter your date of birth in the format DD/MM/YYYY?")
        bot.register_next_step_handler(message, get_dob,name)

def get_register_number(message,name):
    register_number = message.text
    if register_number.isdigit():
        ws.cell(row=ws.max_row,column=4).value = register_number
        bot.send_message(message.chat.id,f"🆔 Thanks! And what is your ID card number,{name}?")
        bot.register_next_step_handler(message,get_id_card,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look like a number. Could you please enter your register number using only digits?")
        bot.register_next_step_handler(message,get_register_number,name)

def get_id_card(message,name):
    id_card = message.text
    if re.match(r'^[a-zA-Z0-9]+$', id_card):
        ws.cell(row=ws.max_row,column=5).value = id_card
        bot.send_message(message.chat.id,f"🏫 Thanks {name}! Which department are you in?")
        bot.register_next_step_handler(message,get_department,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your ID card number using only letters and numbers?")
        bot.register_next_step_handler(message,get_id_card,name)

def get_department(message,name):
    department = message.text
    ws.cell(row=ws.max_row,column=6).value = department
    bot.send_message(message.chat.id,f"📚 Got it! And which year are you in, {name}?")
    bot.register_next_step_handler(message,get_year,name)

def get_year(message,name):
    year = message.text
    if year.isdigit():
        ws.cell(row=ws.max_row,column=7).value = year
        bot.send_message(message.chat.id,f"📱 Thanks {name}! What's your contact number?")
        bot.register_next_step_handler(message,get_contact_number,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look like a number. Could you please enter your year using only digits?")
        bot.register_next_step_handler(message,get_year,name)

def get_contact_number(message,name):
    contact_number = message.text
    if contact_number.startswith('+91') and len(contact_number) == 13:
        ws.cell(row=ws.max_row,column=8).value = contact_number
        bot.send_message(message.chat.id,f"💬 Got it! And what's your WhatsApp number, {name}?")
        bot.register_next_step_handler(message,get_whatsapp_number,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your contact number in the format +91XXXXXXXXXX?")
        bot.register_next_step_handler(message,get_contact_number,name)

def get_whatsapp_number(message,name):
    whatsapp_number = message.text
    if whatsapp_number.startswith('+91') and len(whatsapp_number) == 13:
        ws.cell(row=ws.max_row,column=9).value = whatsapp_number
        bot.send_message(message.chat.id,f"📧 Thanks {name}! What's your email address?")
        bot.register_next_step_handler(message,get_email,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your WhatsApp number in the format +91XXXXXXXXXX?")
        bot.register_next_step_handler(message,get_whatsapp_number,name)

def get_email(message,name):
    email = message.text
    if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
        ws.cell(row=ws.max_row,column=10).value = email
        bot.send_message(message.chat.id,f"👨‍👩‍👧‍👦 Thanks {name}! What's your father's name?")
        bot.register_next_step_handler(message,get_father_name,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter a valid email address?")
        bot.register_next_step_handler(message,get_email,name)

def get_father_name(message,name):
    father_name = message.text
    if re.match(r'^[a-zA-Z\s]+$', father_name):
        ws.cell(row=ws.max_row,column=11).value = father_name
        bot.send_message(message.chat.id,f"👩‍👧‍👦 Got it! And what's your mother's name, {name}?")
        bot.register_next_step_handler(message,get_mother_name,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your father's name using only letters?")
        bot.register_next_step_handler(message,get_father_name,name)

def get_mother_name(message,name):
    mother_name = message.text
    if re.match(r'^[a-zA-Z\s]+$', mother_name):
        ws.cell(row=ws.max_row,column=12).value = mother_name
        bot.send_message(message.chat.id,f"☎️ Thanks {name}! What's your parents' contact number?")
        bot.register_next_step_handler(message,get_parents_number,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your mother's name using only letters?")
        bot.register_next_step_handler(message,get_mother_name,name)

def get_parents_number(message,name):
    parents_number = message.text
    if parents_number.startswith('+91') and len(parents_number) == 13:
        ws.cell(row=ws.max_row,column=13).value = parents_number
        bot.send_message(message.chat.id,f"💼 Got it! What does your father do for a living, {name}?")
        bot.register_next_step_handler(message,get_father_work,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your parents' contact number in the format +91XXXXXXXXXX?")
        bot.register_next_step_handler(message,get_parents_number,name)

def get_father_work(message,name):
    father_work = message.text
    ws.cell(row=ws.max_row,column=14).value = father_work
    bot.send_message(message.chat.id,f"💼 Thanks {name}! And what does your mother do for a living?")
    bot.register_next_step_handler(message,get_mother_work,name)

def get_mother_work(message,name):
    mother_work = message.text
    ws.cell(row=ws.max_row,column=15).value = mother_work
    bot.send_message(message.chat.id,f"📝 Thanks {name}! What was your 10th mark (out of 500)?")
    bot.register_next_step_handler(message,get_10th_mark,name)

def get_10th_mark(message,name):
    mark_10th = message.text
    if mark_10th.isdigit() and int(mark_10th) <= 500:
        ws.cell(row=ws.max_row,column=16).value = mark_10th
        bot.send_message(message.chat.id,f"📝 Got it! And what was your 12th mark (out of 600), {name}?")
        bot.register_next_step_handler(message,get_12th_mark,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your 10th mark as a number out of 500?")
        bot.register_next_step_handler(message,get_10th_mark,name)

def get_12th_mark(message,name):
    mark_12th = message.text
    if mark_12th.isdigit() and int(mark_12th) <= 600:
        ws.cell(row=ws.max_row,column=17).value = mark_12th
        bot.send_message(message.chat.id,f"🏫 Thanks {name}! What was the name of your 10th school?")
        bot.register_next_step_handler(message,get_school_name_10th,name)
    else:
        bot.send_message(message.chat.id,"🤔 That doesn't look quite right. Could you please enter your 12th mark as a number out of 600?")
        bot.register_next_step_handler(message,get_12th_mark,name)

def get_school_name_10th(message,name):
    school_name_10th = message.text
    ws.cell(row=ws.max_row,column=18).value = school_name_10th
    bot.send_message(message.chat.id,f"🏫 Got it! And what was the name of your 12th school, {name}?")
    bot.register_next_step_handler(message,get_school_name_12th,name)

def get_school_name_12th(message,name):
    school_name_12th = message.text
    ws.cell(row=ws.max_row,column=19).value = school_name_12th
    bot.send_message(message.chat.id,f"🏠 Thanks {name}! What's your address?")
    bot.register_next_step_handler(message,get_address,name)

def get_address(message,name):
    address = message.text
    ws.cell(row=ws.max_row,column=20).value = address
    wb.save('user_data.xlsx')

    row = ws[ws.max_row]
    details = f"Name: {row[0].value}\nAge: {row[1].value}\nDOB: {row[2].value}\nRegister Number: {row[3].value}\nID Card: {row[4].value}\nDepartment: {row[5].value}\nYear: {row[6].value}\nContact Number: {row[7].value}\nWhatsApp Number: {row[8].value}\nEmail: {row[9].value}\nFather's Name: {row[10].value}\nMother's Name: {row[11].value}\nParents' Contact Number: {row[12].value}\nFather's Work: {row[13].value}\nMother's Work: {row[14].value}\n10th Mark (out of 500): {row[15].value}\n12th Mark (out of 600): {row[16].value}\n10th School Name: {row[17].value}\n12th School Name: {row[18].value}\nAddress: {row[19].value}"

    bot.send_message(message.chat.id,f"🎉 All done! Here are your details:\n\n{details}")

    global entering_details
    entering_details = False
    notify_waiting_users()


@bot.message_handler(commands=['get_file'])
def get_file(message):
    msg = bot.send_message(message.chat.id, "Please enter the faculty password.")
    bot.register_next_step_handler(msg, send_file)

def send_file(message):
    if message.text == faculty_password:
        with open(filename, 'rb') as file:
            bot.send_document(message.chat.id, file)
    else:
        bot.send_message(message.chat.id, "Incorrect password. Access denied.")

@bot.message_handler(commands=['searchdetails'])
def search_details(message):
    msg = bot.send_message(message.chat.id, "Please enter the faculty password.")
    bot.register_next_step_handler(msg, check_faculty_password)

def check_faculty_password(message):
    if message.text == faculty_password:
        msg = bot.send_message(message.chat.id, "Please enter the name or register number of the user you want to search for.")
        bot.register_next_step_handler(msg, search_user)
    else:
        bot.send_message(message.chat.id, "Incorrect password. Access denied.")


def search_user(message):
    user = message.text
    wb = load_workbook('user_data.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == user or (row[3] and str(row[3]) == user):
            bot.send_message(message.chat.id, f"Name: {row[0]}\nAge: {row[1]}\nDOB: {row[2]}\nRegister Number: {row[3]}\nID Card: {row[4]}\nDepartment: {row[5]}\nYear: {row[6]}\nContact Number: {row[7]}\nWhatsApp Number: {row[8]}\nEmail ID: {row[9]}\nFather Name: {row[10]}\nMother Name: {row[11]}\nParents Number: {row[12]}\nFather Work: {row[13]}\nMother Work: {row[14]}\n10th Mark (out of 500): {row[15]}\n12th Mark (out of 600): {row[16]}\nSchool Name (10th): {row[17]}\nSchool Name (12th): {row[18]}\nAddress: {row[19]}")
            break
    else:
        bot.send_message(message.chat.id, "No user found with that name or register number.")

@bot.message_handler(commands=['editdetails'])
def editdetails(message):
    msg = bot.send_message(message.chat.id, "Please enter the faculty password.")
    bot.register_next_step_handler(msg, edit_details)

def edit_details(message):
    if message.text == faculty_password:
        msg = bot.send_message(message.chat.id, "Please enter the name or register number to edit details for.")
        bot.register_next_step_handler(msg, edit)
    else:
        bot.send_message(message.chat.id, "Incorrect password. Access denied.")

def edit(message):
    search_term = message.text.lower()
    wb = load_workbook(filename)
    ws = wb.active
    found = False
    for row in ws.iter_rows():
        if row[0].value.lower() == search_term or row[18].value.lower() == search_term:
            found = True
            msg = bot.send_message(message.chat.id, "Please enter the new name.")
            bot.register_next_step_handler(msg, edit_name, row)
            break
    if not found:
        bot.send_message(message.chat.id, "No details found for the given name or register number.")

def edit_name(message, row):
    row[0].value = message.text
    msg = bot.send_message(message.chat.id, "Please enter the new age.")
    bot.register_next_step_handler(msg, edit_age,row)

def edit_age(message,row):
    row[1].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter the new date of birth (dd/mm/yyyy).")
    bot.register_next_step_handler(msg,edit_dob,row)

def edit_dob(message,row):
    row[2].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter the new register number.")
    bot.register_next_step_handler(msg,edit_reg_num,row)

def edit_reg_num(message,row):
    row[3].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter the new id card number.")
    bot.register_next_step_handler(msg,edit_id_card,row)

def edit_id_card(message,row):
    row[4].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your new department.")
    bot.register_next_step_handler(msg,edit_department,row)

def edit_department(message,row):
    row[5].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your edited year.")
    bot.register_next_step_handler(msg,edit_year,row)

def edit_year(message,row):
    row[6].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter the new contact number.")
    bot.register_next_step_handler(msg,edit_contact_number,row)

def edit_contact_number(message,row):
    row[7].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your WhatsApp number.")
    bot.register_next_step_handler(msg,edit_whatsapp_number,row)

def edit_whatsapp_number(message,row):
    row[8].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your father name.")
    bot.register_next_step_handler(msg,edit_email,row)

def edit_email(message,row):
    row[9].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter the new email.")
    bot.register_next_step_handler(msg,edit_father_name,row)

def edit_father_name(message,row):
    row[10].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your mother name.")
    bot.register_next_step_handler(msg,edit_mother_name,row)

def edit_mother_name(message,row):
    row[11].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your parents contact number.")
    bot.register_next_step_handler(msg,edit_parents_number,row)

def edit_parents_number(message,row):
    row[12].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your father work.")
    bot.register_next_step_handler(msg,edit_father_work,row)

def edit_father_work(message,row):
    row[13].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your mother work.")
    bot.register_next_step_handler(msg,edit_mother_work,row)


def edit_mother_work(message,row):
    row[14].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your 10th mark.")
    bot.register_next_step_handler(msg,edit_10th_mark,row)

def edit_10th_mark(message,row):
    row[15].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your 12th mark.")
    bot.register_next_step_handler(msg,edit_12th_mark,row)

def edit_12th_mark(message,row):
    row[16].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your 10th school name.")
    bot.register_next_step_handler(msg,edit_10th_school,row)

def edit_10th_school(message,row):
    row[17].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your 12th school name.")
    bot.register_next_step_handler(msg,edit_12th_school,row)

def edit_12th_school(message,row):
    row[18].value=message.text
    msg=bot.send_message(message.chat.id,"Please enter your new address.")
    bot.register_next_step_handler(msg,edit_address,row)

def edit_address(message,row):
    row[19].value=message.text
    wb.save(filename)
    bot.send_message(message.chat.id,"Thank you! The details have been updated.")

@bot.message_handler(commands=['resetsheet'])
def resetsheet(message):
    msg = bot.send_message(message.chat.id, "Please enter the faculty password.")
    bot.register_next_step_handler(msg, reset_sheet)

def reset_sheet(message):
    if message.text == faculty_password:
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Age', 'DOB', 'Register Number', 'ID Card', 'Department', 'Year', 'Contact Number', 'WhatsApp Number', 'Email', 'Father Name', 'Mother Name', 'Parents Number', 'Father Work', 'Mother Work', '10th Mark (out of 500)', '12th Mark (out of 600)', 'School Name (10th)', 'School Name (12th)', 'Address'])
        wb.save(filename)
        bot.send_message(message.chat.id, "The sheet has been reset.")
    else:
        bot.send_message(message.chat.id, "Incorrect password. Access denied.")

@bot.message_handler(commands=['delete_detail'])
def delete_detail(message):
    msg = bot.send_message(message.chat.id, "Please enter the name or register number to delete details for.")
    bot.register_next_step_handler(msg, delete)

def delete(message):
    search_term = message.text.lower()
    wb = load_workbook(filename)
    ws = wb.active
    found = False
    for row in ws.iter_rows():
        if row[0].value.lower() == search_term or row[3].value.lower() == search_term:
            found = True
            # send user details to admin
            user_details = f"Name: {row[0].value}\nAge: {row[1].value}\nDOB: {row[2].value}\nRegister Number: {row[3].value}\nID Card: {row[4].value}\nDepartment: {row[5].value}\nYear: {row[6].value}\nContact Number: {row[7].value}\nWhatsApp Number: {row[8].value}\nEmail ID: {row[9].value}\nFather Name: {row[10].value}\nMother Name: {row[11].value}\nParents Number: {row[12].value}\nFather Work: {row[13].value}\nMother Work: {row[14].value}\n10th Mark (out of 500): {row[15].value}\n12th Mark (out of 600): {row[16].value}\nSchool Name (10th): {row[17].value}\nSchool Name (12th): {row[18].value}\nAddress: {row[19].value}"
            bot.send_message(admin_id, f"User {row[0].value} has deleted their details.\n\n{user_details}")
            # delete user details
            ws.delete_rows(row[0].row)
            wb.save(filename)
            bot.send_message(message.chat.id, "The details have been deleted.")
            break
    if not found:
        bot.send_message(message.chat.id, "No details found for the given name or register number.")



bot_creator = "Boopathi Raja"
bot_creator_profile = "https://t.me/Boopathi_Raj"
bot_purpose = "The purpose of this bot is to help students and faculty of Shanmugha college to store and manage their data in an organized manner."

@bot.message_handler(commands=['feedback'])
def feedback(message):
    keyboard = types.InlineKeyboardMarkup()
    url_button = types.InlineKeyboardButton(text="View Creator's Profile", url=bot_creator_profile)
    keyboard.add(url_button)
    feedback_text = f"‖ CREATOR: {bot_creator} ⚡ ‖\n\n{bot_purpose}\n\nIf you have any feedback about this bot, please send a message to the creator on Telegram."
    bot.send_message(message.chat.id, feedback_text, reply_markup=keyboard)

@bot.message_handler(commands=['help'])
def send_help(message):
    help_text = "*Welcome E U N O I A!*\n This bot helps students and faculty to store and manage their data in an organized manner.\n\n"
    help_text += "Here is a guide on how to use the bot:\n\n"
    help_text += "*Students:*\n"
    help_text += "◾ Use the /enterdetails command to enter your details. The bot will guide you through the process of entering your information.\n\n"
    help_text += "◾ Use the /editdetails command to edit your details. The bot will guide you through the process of updating your information.\n\n"
    help_text += "◾ Use the /delete_detail command to delete your details. The bot will confirm that you want to delete your information before proceeding.\n\n"
    help_text += "◾ Use the /newupdates command to know about the latest updates uploaded by the ADMINS:\n\n\n"
    help_text += "*Faculty:*\n"
    help_text += "◾ Use the /searchdetails command to search for student details. The bot will ask you to enter the name or register number of the student you want to search for.\n\n"
    help_text += "◾ Use the /get_file command to get the excel sheet:\n\n"
    help_text += "◾ Use the /resetsheet command to reset the excel sheet (All stored details will be deleted):\n\n"
    help_text += "◾ Use the /delete_detail command to delete the details in the excel sheet, both students & faculty use this command :\n\n"
    help_text += "◾ Use the /newupdates command to know about the latest updates uploaded by the ADMINS. students and faculty both are use this command:\n\n\n"
    help_text += "*Admins:*\n"
    help_text += "◾ Use the /add_admin command to add a new admin. The bot will ask you to enter the user ID of the user you want to add as an admin.\n\n"
    help_text += "◾ Use the /remove_admin command to remove an admin. The bot will ask you to enter the user ID of the user you want to remove as an admin.\n\n"
    help_text += "◾ Use the /update command to add a new update. The bot will ask you to enter the update text or send a media file with an optional caption.\n\n"
    help_text += "◾ Use the /delete_update command to delete all updates. The bot will confirm that you want to delete all updates before proceeding.\n\n"
    help_text += "If you need further assistance, please don't hesitate to ask for help."
    bot.send_message(message.chat.id, help_text, parse_mode='Markdown')



from telebot import types

@bot.message_handler(commands=['about'])
def send_about(message):
    about_text = "🎓 *Shanmugha College Bot* 🎓\n\n"
    about_text += "This bot was created by a team of talented students as a 3rd year mini project:\n\n"
    about_text += "👨‍💻 *BOOPATHI RAJA. K* (Team leader)\n"
    about_text += "👨‍💻 *KISHORE RAJAN. R*\n"
    about_text += "👨‍💻 *NAVEEN KUMAR. M*\n\n"
    about_text += "Its purpose is to help students and faculty of Shanmugha college to store and manage their data in an organized manner. 📚"

    # Create the inline keyboard markup
    markup = types.InlineKeyboardMarkup()
    btn = types.InlineKeyboardButton("View Creator's Profile", url=bot_creator_profile)
    markup.add(btn)

    # Send the video with the about caption and inline keyboard markup
    video = open('/home/boopathiraj/videos/about_video.mp4', 'rb')
    sent_message = bot.send_video(message.chat.id, video, caption=about_text, reply_markup=markup)

    # Pin the message for 5 seconds
    bot.pin_chat_message(message.chat.id, sent_message.message_id, disable_notification=True)
    time.sleep(10)
    bot.unpin_chat_message(message.chat.id)

@bot.message_handler(commands=['how_it_works'])
def send_how_it_works(message):
    description = """
    Here is a simple User Guide of how to use this bot:\n
    - If you are a student or faculty member at Shanmugha College, you can use the "SSEI" command to access the bot.\n
    - You will be prompted to enter your password, and if it is correct, you will be granted access to use the bot.\n
    - If you are not a student or faculty member at Shanmugha College and are just visiting the bot, you can use the "just_explore" command. However, please note that this will only give you restricted access to the bot’s features.\n
    - The bot has a custom keyboard available with options for STUDENTS, FACULTY, and ADMIN.\n
    - If you click the STUDENTS inline button, you will have access to the student commands.\n
    - If you click the FACULTY inline button, you will have access to the faculty commands.\n
    - If you click the ADMIN inline button, you will have access to the admin commands. However, please note that the admin commands are only available for use by admins.
    """
    bot.send_message(message.chat.id, description)


@bot.message_handler(commands=['Help_Me'])
def send_Help_Me(message):
    description = """
    Here is a description of how to use this bot:\n
    - If you are a student or faculty member at Shanmugha College, you can use the   /SSEI command to access the bot. You will be prompted to enter your password, and if it is correct, you will be granted access to use the bot.\n
    - If you are not a student or faculty member at Shanmugha College and are just visiting the bot, you can use the /just_explore command. However, please note that this will only give you restricted access to the bot’s features.\n
    - If you make a mistake while using the bot and would like to start over, you can either clear the bot’s history or use the /start command to begin a new conversation.
    """
    bot.send_message(message.chat.id, description)



bot.polling()
